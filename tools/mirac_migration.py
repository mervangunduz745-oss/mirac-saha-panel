from __future__ import annotations

import hashlib
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from urllib.request import Request, urlopen


DEFAULT_API_URL = "http://127.0.0.1:8777/api/state"
SEED_FORMAT = "mirac-erp-firestore-subcollections/v1"
COLLECTION_ORDER = (
    "settings",
    "accounts",
    "cariCards",
    "items",
    "recipes",
    "transactions",
    "debtPlans",
    "fixedExpenses",
    "productionJobs",
    "logs",
)

EXCEL_SHEETS = {
    "accounts": "KASA_HESAPLARI",
    "cariCards": "CARI_KARTLARI",
    "items": "STOK_KARTLARI",
    "recipes": "RECETELER",
    "transactions": "HAREKETLER",
    "debtPlans": "BORC_PLANLARI",
    "fixedExpenses": "SABIT_GIDERLER",
    "productionJobs": "URETIM_ISLERI",
    "logs": "LOG",
}

FIELDS = {
    "accounts": ("name", "opening"),
    "cariCards": ("id", "type", "name", "phone", "note", "status", "createdAt", "updatedAt"),
    "items": ("code", "name", "kind", "unit", "opening", "min", "cost"),
    "recipes": ("product", "component", "qty", "note"),
    "transactions": (
        "id", "date", "type", "party", "item", "qty", "unitPrice", "amount", "paid",
        "account", "status", "note", "debtPlanId", "createdAt",
    ),
    "debtPlans": (
        "id", "type", "scope", "party", "amount", "paid", "dueDate", "priority",
        "status", "account", "note", "createdAt",
    ),
    "fixedExpenses": ("id", "name", "category", "amount", "period", "dueDay", "account", "active"),
    "productionJobs": ("id", "customer", "product", "qty", "dueDate", "stage", "priority", "note"),
    "logs": ("at", "event", "detail"),
}

REQUIRED_FIELDS = {
    "accounts": ("name",),
    "cariCards": ("id", "type", "name", "status"),
    "items": ("code", "name", "kind", "unit"),
    "recipes": ("product", "component", "qty"),
    "transactions": ("id", "date", "type", "party", "amount", "status"),
    "debtPlans": ("id", "type", "scope", "party", "amount", "paid", "dueDate", "priority", "status"),
    "fixedExpenses": ("id", "name", "category", "amount", "period", "dueDay", "account", "active"),
    "productionJobs": ("id", "customer", "product", "qty", "dueDate", "stage", "priority"),
    "logs": ("at", "event"),
}

IDENTIFIER_FIELDS = {
    "accounts": ("name",),
    "cariCards": ("id",),
    "items": ("code",),
    "recipes": ("product", "component"),
    "transactions": ("id",),
    "debtPlans": ("id",),
    "fixedExpenses": ("id",),
    "productionJobs": ("id",),
}

NUMBER_FIELDS = {
    "accounts": {"opening"},
    "items": {"opening", "min", "cost"},
    "recipes": {"qty"},
    "transactions": {"qty", "unitPrice", "amount", "paid"},
    "debtPlans": {"amount", "paid"},
    "fixedExpenses": {"amount", "dueDay"},
    "productionJobs": {"qty"},
}

BOOLEAN_FIELDS = {"fixedExpenses": {"active"}}
REFERENCE_TYPES = {"hard", "soft"}


@dataclass(frozen=True)
class Issue:
    severity: str
    category: str
    code: str
    message: str
    collection: str = ""
    row: int | None = None
    field: str = ""
    value: Any = None


@dataclass(frozen=True)
class ReferenceCheck:
    sourceCollection: str
    sourceLegacyId: str
    field: str
    targetCollection: str
    targetLegacyId: str
    referenceType: str
    resolved: bool


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def canonical_json(value: Any) -> bytes:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str).encode("utf-8")


def source_hash(state: dict[str, Any]) -> str:
    return hashlib.sha256(canonical_json(state)).hexdigest()


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _unwrap_state(payload: Any) -> dict[str, Any]:
    if not isinstance(payload, dict):
        raise ValueError("Kaynak JSON nesne olmalı.")
    state = payload.get("state", payload)
    if not isinstance(state, dict):
        raise ValueError("Kaynak içindeki state nesne olmalı.")
    return state


def load_source(source: str, timeout: float = 10.0) -> tuple[dict[str, Any], dict[str, Any], list[Issue]]:
    if source.lower().startswith(("http://", "https://")):
        request = Request(source, method="GET", headers={"Accept": "application/json"})
        with urlopen(request, timeout=timeout) as response:
            if response.status != 200:
                raise ValueError(f"GET kaynağı HTTP {response.status} döndürdü.")
            payload = json.loads(response.read().decode("utf-8-sig"))
        state = _unwrap_state(payload)
        return state, {"kind": "api", "location": source, "readMethod": "GET"}, []

    path = Path(source).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(path)
    if path.suffix.lower() == ".json":
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
        state = _unwrap_state(payload)
        return state, {"kind": "json", "fileName": path.name, "readMethod": "file-read"}, []
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _load_excel(path)
    raise ValueError("Kaynak .json, .xlsx, .xlsm veya HTTP(S) GET adresi olmalı.")


def _load_excel(path: Path) -> tuple[dict[str, Any], dict[str, Any], list[Issue]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel okuma için bundled Python içindeki openpyxl gerekli.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    state: dict[str, Any] = {
        "meta": {"app": "Mirac ERP Pilot", "source": "excel", "workbookFile": path.name}
    }
    issues: list[Issue] = []
    for collection, sheet_name in EXCEL_SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            state[collection] = []
            issues.append(Issue("warning", "source", "MISSING_SHEET", f"Excel sayfası bulunamadı: {sheet_name}", collection))
            continue
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        raw_headers = next(rows, ())
        headers = [str(value).strip() if value is not None else "" for value in raw_headers]
        expected = set(FIELDS[collection])
        for field in sorted(expected - set(headers)):
            issues.append(Issue("warning", "source", "MISSING_COLUMN", f"Excel kolonu bulunamadı: {field}", collection, field=field))
        records = []
        for row_number, values in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in values):
                continue
            record = {
                header: _json_value(values[index]) if index < len(values) else ""
                for index, header in enumerate(headers)
                if header
            }
            records.append(record)
        state[collection] = records
    workbook.close()
    return state, {"kind": "excel", "fileName": path.name, "readMethod": "read-only"}, issues


def _parse_number(value: Any) -> int | float | Any:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not math.isfinite(value):
            return value
        return int(value) if float(value).is_integer() else value
    if not isinstance(value, str):
        return value
    text = value.strip()
    if text == "":
        return 0
    compact = text.replace(" ", "")
    if "," in compact and "." in compact:
        compact = compact.replace(".", "").replace(",", ".")
    elif "," in compact:
        compact = compact.replace(",", ".")
    try:
        parsed = float(compact)
    except ValueError:
        return value
    return int(parsed) if parsed.is_integer() else parsed


def _parse_boolean(value: Any) -> bool | Any:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in {0, 1}:
        return bool(value)
    if isinstance(value, str):
        folded = value.strip().casefold()
        if folded in {"true", "1", "evet", "aktif", "yes"}:
            return True
        if folded in {"false", "0", "hayır", "hayir", "pasif", "no"}:
            return False
    return value


def _clean_scalar(value: Any) -> Any:
    value = _json_value(value)
    if isinstance(value, str):
        return value.strip()
    return value


def normalize_state(raw_state: dict[str, Any]) -> tuple[dict[str, Any], list[Issue]]:
    issues: list[Issue] = []
    raw_meta = raw_state.get("meta", {})
    meta = dict(raw_meta) if isinstance(raw_meta, dict) else {}
    meta.pop("workbook", None)
    meta.pop("workbookFile", None)
    normalized: dict[str, Any] = {"meta": {key: _clean_scalar(value) for key, value in meta.items()}}

    for collection in FIELDS:
        records = raw_state.get(collection, [])
        if not isinstance(records, list):
            normalized[collection] = []
            issues.append(Issue("error", "schema", "COLLECTION_NOT_LIST", "Koleksiyon liste olmalı.", collection, value=type(records).__name__))
            continue
        clean_records = []
        for row_number, record in enumerate(records, start=1):
            if not isinstance(record, dict):
                issues.append(Issue("error", "schema", "RECORD_NOT_OBJECT", "Kayıt nesne olmalı.", collection, row_number, value=type(record).__name__))
                continue
            clean: dict[str, Any] = {}
            for key, value in record.items():
                value = _clean_scalar(value)
                if key in NUMBER_FIELDS.get(collection, set()):
                    value = _parse_number(value)
                elif key in BOOLEAN_FIELDS.get(collection, set()):
                    value = _parse_boolean(value)
                clean[str(key)] = value
            for field in FIELDS[collection]:
                if field not in clean:
                    if field in NUMBER_FIELDS.get(collection, set()):
                        clean[field] = 0
                    elif field in BOOLEAN_FIELDS.get(collection, set()):
                        clean[field] = False
                    else:
                        clean[field] = ""
            clean_records.append(clean)
        normalized[collection] = clean_records
    return normalized, issues


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _key(value: Any) -> str:
    return str(value).strip().casefold()


def legacy_id(collection: str, record: dict[str, Any], row_number: int) -> str:
    if collection == "accounts":
        return str(record.get("name", "")).strip()
    if collection == "items":
        return str(record.get("code", "")).strip()
    if collection == "recipes":
        return f"{record.get('product', '')}|{record.get('component', '')}"
    if collection == "logs":
        material = f"{record.get('at', '')}|{record.get('event', '')}|{row_number}"
        return f"LOG-{hashlib.sha256(material.encode('utf-8')).hexdigest()[:16]}"
    return str(record.get("id", "")).strip()


def _ascii_slug(value: str) -> str:
    replacements = str.maketrans({"ı": "i", "İ": "I", "ş": "s", "Ş": "S", "ğ": "g", "Ğ": "G", "ü": "u", "Ü": "U", "ö": "o", "Ö": "O", "ç": "c", "Ç": "C"})
    normalized = unicodedata.normalize("NFKD", value.translate(replacements)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^A-Za-z0-9_-]+", "-", normalized).strip("-_").lower()


def document_id(collection: str, legacy: str) -> str:
    if collection in {"cariCards", "items", "transactions", "debtPlans", "fixedExpenses", "productionJobs"}:
        candidate = legacy
    else:
        candidate = _ascii_slug(legacy)
    candidate = candidate.replace("/", "-").strip()
    invalid = not candidate or candidate in {".", ".."} or re.fullmatch(r"__.*__", candidate) or len(candidate.encode("utf-8")) > 1400
    if invalid:
        candidate = f"{collection[:8]}-{hashlib.sha256(legacy.encode('utf-8')).hexdigest()[:20]}"
    return candidate


def validate_state(state: dict[str, Any]) -> tuple[list[Issue], list[ReferenceCheck]]:
    issues: list[Issue] = []
    for collection in FIELDS:
        records = state.get(collection)
        if not isinstance(records, list):
            issues.append(Issue("error", "schema", "COLLECTION_NOT_LIST", "Koleksiyon liste olmalı.", collection))
            continue
        seen: dict[tuple[str, ...], list[int]] = defaultdict(list)
        id_fields = IDENTIFIER_FIELDS.get(collection)
        for row_number, record in enumerate(records, start=1):
            if not isinstance(record, dict):
                issues.append(Issue("error", "schema", "RECORD_NOT_OBJECT", "Kayıt nesne olmalı.", collection, row_number))
                continue
            for field in REQUIRED_FIELDS[collection]:
                value = record.get(field)
                if _is_blank(value):
                    issues.append(Issue("error", "schema", "REQUIRED_FIELD_MISSING", "Zorunlu alan boş.", collection, row_number, field, value))
            for field in NUMBER_FIELDS.get(collection, set()):
                value = record.get(field)
                if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(float(value)):
                    issues.append(Issue("error", "schema", "INVALID_NUMBER", "Alan sonlu sayı olmalı.", collection, row_number, field, value))
            for field in BOOLEAN_FIELDS.get(collection, set()):
                value = record.get(field)
                if not isinstance(value, bool):
                    issues.append(Issue("error", "schema", "INVALID_BOOLEAN", "Alan boolean olmalı.", collection, row_number, field, value))
            if id_fields and not any(_is_blank(record.get(field)) for field in id_fields):
                seen[tuple(_key(record.get(field)) for field in id_fields)].append(row_number)
        for duplicate_key, rows in seen.items():
            if len(rows) > 1:
                issues.append(Issue(
                    "error", "collision", "DUPLICATE_LEGACY_IDENTIFIER",
                    f"Tekrarlı kimlik/kod: {' | '.join(duplicate_key)}; satırlar {rows}.", collection,
                    field="+".join(id_fields or ()), value={"rows": rows, "key": duplicate_key},
                ))

    issues.extend(_validate_debt_plans(state))
    references = _check_references(state)
    for ref in references:
        if not ref.resolved:
            severity = "error" if ref.referenceType == "hard" else "warning"
            issues.append(Issue(
                severity, "reference", "UNRESOLVED_REFERENCE",
                f"{ref.targetCollection} hedefi bulunamadı: {ref.targetLegacyId}",
                ref.sourceCollection, field=ref.field,
                value={"sourceLegacyId": ref.sourceLegacyId, "targetLegacyId": ref.targetLegacyId, "referenceType": ref.referenceType},
            ))
    return issues, references


def _validate_debt_plans(state: dict[str, Any]) -> list[Issue]:
    issues: list[Issue] = []
    plans = state.get("debtPlans", []) if isinstance(state.get("debtPlans"), list) else []
    duplicate_candidates: dict[tuple[str, str, float, str], list[str]] = defaultdict(list)
    for row_number, plan in enumerate(plans, start=1):
        if not isinstance(plan, dict):
            continue
        plan_id = str(plan.get("id", ""))
        amount = plan.get("amount")
        paid = plan.get("paid")
        if isinstance(amount, (int, float)) and not isinstance(amount, bool):
            if amount <= 0:
                issues.append(Issue("error", "debt", "DEBT_AMOUNT_NOT_POSITIVE", "Borç tutarı 0'dan büyük olmalı.", "debtPlans", row_number, "amount", amount))
        if isinstance(paid, (int, float)) and not isinstance(paid, bool):
            if paid < 0:
                issues.append(Issue("error", "debt", "DEBT_PAID_NEGATIVE", "Ödenen tutar negatif olamaz.", "debtPlans", row_number, "paid", paid))
            if isinstance(amount, (int, float)) and paid > amount:
                issues.append(Issue("error", "debt", "DEBT_OVERPAID", "Ödenen tutar toplam borçtan büyük.", "debtPlans", row_number, "paid", {"amount": amount, "paid": paid}))
            remaining = float(amount) - float(paid) if isinstance(amount, (int, float)) else None
            status = str(plan.get("status", ""))
            if status == "Ödendi" and remaining is not None and remaining > 0.000001:
                issues.append(Issue("error", "debt", "DEBT_CLOSED_WITH_BALANCE", f"Ödendi durumunda {remaining:g} kalan var.", "debtPlans", row_number, "status", plan_id))
            if status == "Kısmi" and (paid <= 0 or (remaining is not None and remaining <= 0)):
                issues.append(Issue("error", "debt", "DEBT_PARTIAL_STATUS_MISMATCH", "Kısmi durumu ödenen/kalan tutarla uyuşmuyor.", "debtPlans", row_number, "status", plan_id))
        if isinstance(amount, (int, float)):
            candidate = (_key(plan.get("party")), _key(plan.get("type")), float(amount), str(plan.get("dueDate", "")))
            duplicate_candidates[candidate].append(plan_id)
    for candidate, ids in duplicate_candidates.items():
        if len(ids) > 1:
            issues.append(Issue("warning", "debt", "POSSIBLE_DUPLICATE_DEBT_PLAN", f"Aynı kişi/tür/tutar/vadede olası tekrar: {ids}.", "debtPlans", value={"ids": ids, "key": candidate}))

    plan_ids = {_key(plan.get("id")) for plan in plans if isinstance(plan, dict) and not _is_blank(plan.get("id"))}
    for row_number, tx in enumerate(state.get("transactions", []), start=1):
        if not isinstance(tx, dict):
            continue
        tx_type = str(tx.get("type", ""))
        ref = str(tx.get("debtPlanId", "")).strip()
        if tx_type == "BORC_ODEME" and not ref:
            issues.append(Issue("error", "debt", "DEBT_PAYMENT_REFERENCE_MISSING", "Borç ödeme hareketinde debtPlanId boş.", "transactions", row_number, "debtPlanId", tx.get("id")))
        if ref and _key(ref) not in plan_ids:
            looks_like_timestamp = bool(re.match(r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}", ref))
            code = "SUSPECT_CREATED_AT_IN_DEBT_PLAN_ID" if looks_like_timestamp and not tx.get("createdAt") else "DANGLING_DEBT_PLAN_ID"
            issues.append(Issue("error" if tx_type == "BORC_ODEME" else "warning", "debt", code, "debtPlanId hiçbir borç planına bağlanmıyor.", "transactions", row_number, "debtPlanId", {"transactionId": tx.get("id"), "debtPlanId": ref}))
    return issues


def _check_references(state: dict[str, Any]) -> list[ReferenceCheck]:
    indexes = {
        "accounts": {_key(row.get("name")) for row in state.get("accounts", []) if isinstance(row, dict)},
        "cariCards": {_key(row.get("name")) for row in state.get("cariCards", []) if isinstance(row, dict)},
        "items": {_key(row.get("code")) for row in state.get("items", []) if isinstance(row, dict)},
        "debtPlans": {_key(row.get("id")) for row in state.get("debtPlans", []) if isinstance(row, dict)},
    }
    refs: list[ReferenceCheck] = []

    def add(source_collection: str, record: dict[str, Any], row_number: int, field: str, target: str, kind: str) -> None:
        value = str(record.get(field, "")).strip()
        if not value:
            return
        refs.append(ReferenceCheck(source_collection, legacy_id(source_collection, record, row_number), field, target, value, kind, _key(value) in indexes[target]))

    for row_number, row in enumerate(state.get("recipes", []), start=1):
        if isinstance(row, dict):
            add("recipes", row, row_number, "product", "items", "hard")
            add("recipes", row, row_number, "component", "items", "hard")
    for row_number, row in enumerate(state.get("transactions", []), start=1):
        if not isinstance(row, dict):
            continue
        if row.get("item"):
            add("transactions", row, row_number, "item", "items", "hard")
        if row.get("account"):
            add("transactions", row, row_number, "account", "accounts", "hard")
        if row.get("type") == "BORC_ODEME" and row.get("debtPlanId"):
            add("transactions", row, row_number, "debtPlanId", "debtPlans", "hard")
        if row.get("party") and row.get("party") not in {"Üretim", "Tanımsız"}:
            add("transactions", row, row_number, "party", "cariCards", "soft")
    for row_number, row in enumerate(state.get("debtPlans", []), start=1):
        if isinstance(row, dict):
            add("debtPlans", row, row_number, "account", "accounts", "hard")
            add("debtPlans", row, row_number, "party", "cariCards", "soft")
    for row_number, row in enumerate(state.get("fixedExpenses", []), start=1):
        if isinstance(row, dict):
            add("fixedExpenses", row, row_number, "account", "accounts", "hard")
    for row_number, row in enumerate(state.get("productionJobs", []), start=1):
        if isinstance(row, dict):
            add("productionJobs", row, row_number, "product", "items", "hard")
            add("productionJobs", row, row_number, "customer", "cariCards", "soft")
    return refs


def _settings_record(state: dict[str, Any]) -> dict[str, Any]:
    return dict(state.get("meta", {})) if isinstance(state.get("meta"), dict) else {}


def build_seed_bundle(
    state: dict[str, Any], org_id: str, schema_version: int, import_batch_id: str, source: dict[str, Any]
) -> tuple[dict[str, Any], list[Issue]]:
    if not re.fullmatch(r"[A-Za-z0-9_-]{2,120}", org_id):
        raise ValueError("orgId yalnızca harf, rakam, _ ve - içermeli; 2-120 karakter olmalı.")
    if schema_version < 1:
        raise ValueError("schemaVersion en az 1 olmalı.")
    if not re.fullmatch(r"[A-Za-z0-9._-]{3,160}", import_batch_id):
        raise ValueError("importBatchId güvenli karakterlerden oluşmalı; 3-160 karakter olmalı.")

    collections: dict[str, dict[str, Any]] = {}
    collision_issues: list[Issue] = []
    for collection in COLLECTION_ORDER:
        records = [_settings_record(state)] if collection == "settings" else state.get(collection, [])
        documents = []
        seen_doc_ids: dict[str, list[str]] = defaultdict(list)
        for row_number, record in enumerate(records, start=1):
            if not isinstance(record, dict):
                continue
            source_id = "meta" if collection == "settings" else legacy_id(collection, record, row_number)
            doc_id = "app" if collection == "settings" else document_id(collection, source_id)
            seen_doc_ids[doc_id].append(source_id)
            data = dict(record)
            data.update({"legacyId": source_id, "schemaVersion": schema_version, "importBatchId": import_batch_id})
            documents.append({
                "id": doc_id,
                "path": f"orgs/{org_id}/{collection}/{doc_id}",
                "data": data,
            })
        for doc_id, source_ids in seen_doc_ids.items():
            if len(source_ids) > 1:
                collision_issues.append(Issue(
                    "error", "collision", "DOCUMENT_ID_COLLISION",
                    f"Üretilen belge ID çakışması: {doc_id}", collection,
                    value={"documentId": doc_id, "legacyIds": source_ids},
                ))
        collections[collection] = {
            "collectionPath": f"orgs/{org_id}/{collection}",
            "documents": documents,
        }

    bundle = {
        "format": SEED_FORMAT,
        "orgId": org_id,
        "schemaVersion": schema_version,
        "importBatchId": import_batch_id,
        "generatedAt": utc_now(),
        "source": source,
        "sourceSha256": source_hash(state),
        "collections": collections,
    }
    return bundle, collision_issues


def build_import_batch_id(state: dict[str, Any]) -> str:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return f"mirac-{stamp}-{source_hash(state)[:8]}"


def make_report(
    state: dict[str, Any], bundle: dict[str, Any], issues: list[Issue], references: list[ReferenceCheck]
) -> dict[str, Any]:
    severity_counts = Counter(issue.severity for issue in issues)
    category_counts = Counter(issue.category for issue in issues)
    structural_errors = [issue for issue in issues if issue.severity == "error" and issue.category in {"schema", "collision"}]
    business_errors = [issue for issue in issues if issue.severity == "error" and issue.category not in {"schema", "collision"}]
    unresolved = [ref for ref in references if not ref.resolved]
    collection_counts = {
        name: len(payload["documents"])
        for name, payload in bundle["collections"].items()
    }
    return {
        "format": "mirac-erp-migration-dry-run/v1",
        "generatedAt": bundle["generatedAt"],
        "orgId": bundle["orgId"],
        "schemaVersion": bundle["schemaVersion"],
        "importBatchId": bundle["importBatchId"],
        "source": bundle["source"],
        "sourceSha256": bundle["sourceSha256"],
        "summary": {
            "schemaValid": not structural_errors,
            "uploadReady": not structural_errors and not business_errors,
            "reviewRequired": bool(issues),
            "seedDocumentCount": sum(collection_counts.values()),
            "collectionCounts": collection_counts,
            "issueCounts": {"error": severity_counts["error"], "warning": severity_counts["warning"], "info": severity_counts["info"]},
            "categoryCounts": dict(sorted(category_counts.items())),
            "unresolvedReferenceCount": len(unresolved),
        },
        "collisions": [asdict(issue) for issue in issues if issue.category == "collision"],
        "requiredFieldIssues": [asdict(issue) for issue in issues if issue.code == "REQUIRED_FIELD_MISSING"],
        "debtPlanIssues": [asdict(issue) for issue in issues if issue.category == "debt"],
        "references": [asdict(ref) for ref in references],
        "unresolvedReferences": [asdict(ref) for ref in unresolved],
        "issues": [asdict(issue) for issue in issues],
    }


def report_markdown(report: dict[str, Any]) -> str:
    summary = report["summary"]
    lines = [
        "# Miraç ERP Firebase Migrasyon Dry-Run",
        "",
        f"- Org: `{report['orgId']}`",
        f"- Import batch: `{report['importBatchId']}`",
        f"- Schema version: `{report['schemaVersion']}`",
        f"- Şema geçerli: `{'EVET' if summary['schemaValid'] else 'HAYIR'}`",
        f"- Yüklemeye hazır: `{'EVET' if summary['uploadReady'] else 'HAYIR'}`",
        f"- Belge sayısı: `{summary['seedDocumentCount']}`",
        f"- Hata / uyarı: `{summary['issueCounts']['error']} / {summary['issueCounts']['warning']}`",
        "",
        "## Koleksiyonlar",
        "",
        "| Yol | Belge |",
        "|---|---:|",
    ]
    for collection, count in summary["collectionCounts"].items():
        lines.append(f"| `orgs/{report['orgId']}/{collection}` | {count} |")

    sections = (
        ("ID ve Kod Çakışmaları", report["collisions"]),
        ("Zorunlu Alan Sorunları", report["requiredFieldIssues"]),
        ("Borç Planı Tutarsızlıkları", report["debtPlanIssues"]),
        ("Çözülmeyen Referanslar", report["unresolvedReferences"]),
    )
    for title, rows in sections:
        lines.extend(["", f"## {title}", ""])
        if not rows:
            lines.append("Yok.")
            continue
        for row in rows:
            if "sourceCollection" in row:
                lines.append(
                    f"- `{row['sourceCollection']}/{row['sourceLegacyId']}` `{row['field']}` -> "
                    f"`{row['targetCollection']}/{row['targetLegacyId']}` ({row['referenceType']})"
                )
            else:
                location = row.get("collection", "")
                if row.get("row") is not None:
                    location += f" satır {row['row']}"
                lines.append(f"- `{row['code']}` {location}: {row['message']}")

    lines.extend([
        "",
        "## Güvenlik",
        "",
        "Bu çıktı yalnızca GET/dosya okuma ve yerel JSON üretimi yapar. Firebase'e yazma veya yükleme işlemi içermez.",
        "",
    ])
    return "\n".join(lines)


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
    temp.replace(path)


def write_seed_bundle(bundle: dict[str, Any], output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    existing = [path for path in output_dir.iterdir() if path.is_file()]
    if existing:
        raise FileExistsError(f"Çıktı klasörü boş değil: {output_dir}")
    written: list[Path] = []
    manifest_collections = []
    for collection in COLLECTION_ORDER:
        payload = bundle["collections"][collection]
        file_name = f"{collection}.json"
        write_json(output_dir / file_name, payload)
        written.append(output_dir / file_name)
        manifest_collections.append({
            "name": collection,
            "collectionPath": payload["collectionPath"],
            "file": file_name,
            "documentCount": len(payload["documents"]),
        })
    manifest = {key: value for key, value in bundle.items() if key != "collections"}
    manifest["collections"] = manifest_collections
    write_json(output_dir / "manifest.json", manifest)
    written.append(output_dir / "manifest.json")
    return written


def run_migration(
    source: str, org_id: str, schema_version: int, import_batch_id: str | None = None, timeout: float = 10.0
) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    raw_state, source_meta, source_issues = load_source(source, timeout=timeout)
    state, normalization_issues = normalize_state(raw_state)
    validation_issues, references = validate_state(state)
    batch_id = import_batch_id or build_import_batch_id(state)
    source_meta = {**source_meta, "readOnly": True}
    bundle, collision_issues = build_seed_bundle(state, org_id, schema_version, batch_id, source_meta)
    issues = source_issues + normalization_issues + validation_issues + collision_issues
    report = make_report(state, bundle, issues, references)
    return state, bundle, report


def validate_seed_files(seed_dir: Path) -> list[Issue]:
    issues: list[Issue] = []
    manifest_path = seed_dir / "manifest.json"
    if not manifest_path.exists():
        return [Issue("error", "schema", "MANIFEST_MISSING", "manifest.json bulunamadı.")]
    manifest = json.loads(manifest_path.read_text(encoding="utf-8-sig"))
    if manifest.get("format") != SEED_FORMAT:
        issues.append(Issue("error", "schema", "INVALID_SEED_FORMAT", "Seed formatı desteklenmiyor.", value=manifest.get("format")))
    expected = {entry.get("name"): entry for entry in manifest.get("collections", []) if isinstance(entry, dict)}
    all_paths: dict[str, list[str]] = defaultdict(list)
    for collection in COLLECTION_ORDER:
        entry = expected.get(collection)
        if not entry:
            issues.append(Issue("error", "schema", "COLLECTION_FILE_MISSING", "Manifest koleksiyon kaydı eksik.", collection))
            continue
        path = seed_dir / str(entry.get("file", ""))
        if not path.exists():
            issues.append(Issue("error", "schema", "COLLECTION_FILE_MISSING", f"Dosya bulunamadı: {path.name}", collection))
            continue
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
        documents = payload.get("documents")
        if not isinstance(documents, list):
            issues.append(Issue("error", "schema", "DOCUMENTS_NOT_LIST", "documents liste olmalı.", collection))
            continue
        for row_number, document in enumerate(documents, start=1):
            if not isinstance(document, dict) or not isinstance(document.get("data"), dict):
                issues.append(Issue("error", "schema", "INVALID_SEED_DOCUMENT", "Belge ve data nesne olmalı.", collection, row_number))
                continue
            path_value = str(document.get("path", ""))
            all_paths[path_value].append(str(document.get("data", {}).get("legacyId", "")))
            for field in ("legacyId", "schemaVersion", "importBatchId"):
                if _is_blank(document["data"].get(field)):
                    issues.append(Issue("error", "schema", "MIGRATION_FIELD_MISSING", "Migrasyon alanı eksik.", collection, row_number, field))
            if document["data"].get("schemaVersion") != manifest.get("schemaVersion"):
                issues.append(Issue("error", "schema", "SCHEMA_VERSION_MISMATCH", "Belge schemaVersion manifest ile uyuşmuyor.", collection, row_number, "schemaVersion"))
            if document["data"].get("importBatchId") != manifest.get("importBatchId"):
                issues.append(Issue("error", "schema", "IMPORT_BATCH_MISMATCH", "Belge importBatchId manifest ile uyuşmuyor.", collection, row_number, "importBatchId"))
    for path_value, legacy_ids in all_paths.items():
        if len(legacy_ids) > 1:
            issues.append(Issue("error", "collision", "DOCUMENT_PATH_COLLISION", f"Tekrarlı belge yolu: {path_value}", value={"legacyIds": legacy_ids}))
    return issues
